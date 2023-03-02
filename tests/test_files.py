import os
import zipfile
import csv
from os.path import basename
from openpyxl import load_workbook
from PyPDF2 import PdfReader


path_to_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../files')
path_to_res = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../res')
files_dir = os.listdir(path_to_files)
path_zip = os.path.join(path_to_res, 'file_archive.zip')


def test_create_archive():
    with zipfile.ZipFile(path_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in files_dir:
            add_file = os.path.join(path_to_files, file)
            zf.write(add_file, basename(add_file))


def test_csv():
    with zipfile.ZipFile(path_zip) as myzip:
        csv_doc = myzip.extract('personal_data.csv')
        with open(csv_doc) as csv_file:
            table = csv.reader(csv_file, delimiter=';')
            for line_num, line in enumerate(table, 2):
                if line_num == 1:
                    assert 'Fox' in line[1]
        os.remove(csv_doc)


def test_xlsx():
    with zipfile.ZipFile(path_zip) as myzip:
        xlsx_doc = myzip.extract('x-file.xlsx')
        workbook = load_workbook(xlsx_doc)
        sheet = workbook.active
        sheet = sheet.cell(row=2, column=2).value
        assert 'Samantha' in sheet
        os.remove(xlsx_doc)


def test_pdf():
    with zipfile.ZipFile(path_zip) as myzip:
        pdf_doc = myzip.extract('about the truth.pdf')
        pdf_data = PdfReader(pdf_doc)
        page = pdf_data.pages[0]
        text = page.extract_text()
        assert 'out there' in text
        os.remove(pdf_doc)
